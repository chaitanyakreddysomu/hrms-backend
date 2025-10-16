# utils.py
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase import pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os
from datetime import datetime, date
import calendar
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from django.http import HttpResponse

# Register Bookman Old Style font (as per requirements)
def register_fonts():
    """Register custom fonts for PDF generation"""
    try:
        font_path = os.path.join(settings.BASE_DIR, 'fonts', 'bookman-old-style.ttf')
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('BookmanOldStyle', font_path))
    except:
        pass  # Fallback to default fonts

def generate_pdf_report(template_name, context, filename):
    """
    Generate PDF report using ReportLab
    """
    register_fonts()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        fontName='BookmanOldStyle',
        alignment=1,  # Center alignment
        spaceAfter=30,
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        fontName='BookmanOldStyle',
    )
    
    story = []
    
    if template_name == 'payslip':
        story = generate_payslip_content(context, title_style, normal_style)
    elif template_name == 'attendance_report':
        story = generate_attendance_report_content(context, title_style, normal_style)
    elif template_name == 'salary_statement':
        story = generate_salary_statement_content(context, title_style, normal_style)
    elif template_name == 'invoice':
        story = generate_invoice_content(context, title_style, normal_style)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_payslip_content(context, title_style, normal_style):
    """Generate payslip PDF content"""
    story = []
    
    # Title
    title = Paragraph(f"PAY SLIP - {context['month']}/{context['year']}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Company header
    company_info = [
        ['Company:', context['company_name']],
        ['Employee:', f"{context['employee_name']} ({context['employee_code']})"],
        ['Department:', context['department']],
        ['Designation:', context['designation']],
        ['Date of Joining:', context['date_of_joining']],
    ]
    
    header_table = Table(company_info, colWidths=[2*inch, 4*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(header_table)
    story.append(Spacer(1, 12))
    
    # Salary details
    salary_data = [
        ['EARNINGS', 'AMOUNT', 'DEDUCTIONS', 'AMOUNT'],
        ['Basic', f"₹{context['basic']}", 'PF', f"₹{context['pf_deduction']}"],
        ['DA', f"₹{context['da']}", 'ESI', f"₹{context['esi_deduction']}"],
        ['HRA', f"₹{context['hra']}", 'PT', f"₹{context['pt_deduction']}"],
        ['Conveyance', f"₹{context['conveyance']}", 'LWF', f"₹{context['lwf_deduction']}"],
        ['Bonus', f"₹{context['bonus']}", 'Advance', f"₹{context['advance']}"],
        ['Other Allowances', f"₹{context['other_allowances']}", 'Insurance', f"₹{context['insurance']}"],
        ['GROSS EARNINGS', f"₹{context['gross_salary']}", 'TOTAL DEDUCTIONS', f"₹{context['total_deductions']}"],
        ['', '', 'NET SALARY', f"₹{context['net_salary']}"],
    ]
    
    salary_table = Table(salary_data, colWidths=[2*inch, 1.5*inch, 2*inch, 1.5*inch])
    salary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('BACKGROUND', (0, -2), (-1, -2), colors.lightgrey),
        ('BACKGROUND', (2, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'BookmanOldStyle'),
        ('FONTNAME', (0, -2), (-1, -1), 'BookmanOldStyle'),
    ]))
    
    story.append(salary_table)
    story.append(Spacer(1, 12))
    
    # Footer
    footer_text = f"Net Salary in Words: {context['net_salary_words']}"
    footer = Paragraph(footer_text, normal_style)
    story.append(footer)
    
    return story

def generate_attendance_report_content(context, title_style, normal_style):
    """Generate attendance report PDF content"""
    story = []
    
    title = Paragraph(f"ATTENDANCE STATEMENT - {context['month']}/{context['year']}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Company header
    company_header = Paragraph(f"{context['company_name']}<br/>Deputed at {context['client_name']} - {context['location']}", normal_style)
    story.append(company_header)
    story.append(Spacer(1, 12))
    
    # Attendance data
    headers = ['Sl No', 'ID No', 'Name', 'Designation', 'Department', 'DOJ']
    
    # Add day columns
    days_in_month = context['days_in_month']
    for day in range(1, days_in_month + 1):
        headers.append(str(day))
    
    headers.extend(['Working Days', 'Holidays', 'Weekly Off', 'Absent', 'Total Days', 'OT Hours'])
    
    attendance_data = [headers]
    
    for idx, employee in enumerate(context['employees'], 1):
        row = [
            str(idx),
            employee['employee_code'],
            employee['name'],
            employee['designation'],
            employee['department'],
            employee['doj']
        ]
        
        # Add attendance for each day
        for day in range(1, days_in_month + 1):
            attendance_status = employee['attendance'].get(str(day), 'A')
            row.append(attendance_status)
        
        row.extend([
            str(employee['working_days']),
            str(employee['holidays']),
            str(employee['weekly_offs']),
            str(employee['absent']),
            str(employee['total_days']),
            str(employee['ot_hours'])
        ])
        
        attendance_data.append(row)
    
    attendance_table = Table(attendance_data)
    attendance_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ]))
    
    story.append(attendance_table)
    
    return story

def generate_salary_statement_content(context, title_style, normal_style):
    """Generate salary statement PDF content"""
    story = []
    
    title = Paragraph(f"SALARY STATEMENT - {context['month']}/{context['year']}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Company info
    company_info = Paragraph(f"{context['company_name']}<br/>Employees deputed at {context['client_name']}", normal_style)
    story.append(company_info)
    story.append(Spacer(1, 12))
    
    # Salary details table
    headers = [
        'Sl No', 'Emp Code', 'Name', 'Gender', 'Designation', 'Department',
        'DOB', 'ESI No', 'UAN No', 'DOJ', 'Basic', 'DA', 'Special Allow',
        'Leave with wages', 'Bonus', 'Gross Salary', 'Days Payable',
        'ESI Ded', 'PF Ded', 'PT', 'Total Ded', 'Take Home'
    ]
    
    salary_data = [headers]
    
    for idx, employee in enumerate(context['employees'], 1):
        row = [
            str(idx),
            employee['employee_code'],
            employee['name'],
            employee['gender'],
            employee['designation'],
            employee['department'],
            employee['dob'],
            employee['esi_no'],
            employee['uan_no'],
            employee['doj'],
            f"₹{employee['basic']}",
            f"₹{employee['da']}",
            f"₹{employee['special_allowance']}",
            f"₹{employee['leave_with_wages']}",
            f"₹{employee['bonus']}",
            f"₹{employee['gross_salary']}",
            str(employee['days_payable']),
            f"₹{employee['esi_deduction']}",
            f"₹{employee['pf_deduction']}",
            f"₹{employee['pt_deduction']}",
            f"₹{employee['total_deductions']}",
            f"₹{employee['take_home']}"
        ]
        salary_data.append(row)
    
    salary_table = Table(salary_data)
    salary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ]))
    
    story.append(salary_table)
    return story

def generate_invoice_content(context, title_style, normal_style):
    """Generate invoice PDF content"""
    story = []
    
    title = Paragraph("TAX INVOICE", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Invoice header with company details
    header_data = [
        [context['from_company'], context['to_company']],
        [f"GST No: {context['from_gst']}", f"GST No: {context['to_gst']}"]
    ]
    
    header_table = Table(header_data, colWidths=[3*inch, 3*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(header_table)
    story.append(Spacer(1, 12))
    
    # Invoice details
    invoice_info = [
        ['Invoice No:', context['invoice_no'], 'Invoice Date:', context['invoice_date']],
    ]
    
    info_table = Table(invoice_info, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 12))
    
    # Invoice items
    items_data = [
        ['Sl. No.', 'Description of Services', 'HSN/SAC Code', 'Amount'],
        ['1', f"Contract Labour charges for {context['month']}/{context['year']}", '998519', f"₹{context['gross_amount']}"],
        ['2', 'PF Employer Contribution @ 13%', '998519', f"₹{context['pf_employer']}"],
        ['3', 'ESI Employer Contribution @ 3.25%', '998519', f"₹{context['esi_employer']}"],
        ['4', 'Service Charges', '998519', f"₹{context['service_charge']}"],
        ['', 'Total', '', f"₹{context['subtotal']}"],
        ['', 'Add: CGST @ 9%', '', f"₹{context['cgst']}"],
        ['', 'Add: SGST @ 9%', '', f"₹{context['sgst']}"],
        ['', 'Total Amount', '', f"₹{context['total_amount']}"],
    ]
    
    items_table = Table(items_data, colWidths=[0.8*inch, 3*inch, 1.2*inch, 2*inch])
    items_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'BookmanOldStyle'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('BACKGROUND', (0, -3), (-1, -1), colors.lightgrey),
    ]))
    
    story.append(items_table)
    story.append(Spacer(1, 12))
    
    # Total in words
    total_words = Paragraph(f"Total Amount In Words: {context['total_words']}", normal_style)
    story.append(total_words)
    story.append(Spacer(1, 24))
    
    # Bank details
    bank_details = Paragraph(f"Bank Details:<br/>Bank Name: {context['bank_name']}<br/>Account Number: {context['account_number']}<br/>IFSC Code: {context['ifsc_code']}", normal_style)
    story.append(bank_details)
    story.append(Spacer(1, 24))
    
    # Signature
    signature = Paragraph("For " + context['from_company'] + "<br/><br/><br/>Authorised Signatory", normal_style)
    story.append(signature)
    
    return story

def send_email_with_attachment(to_email, subject, message, attachment_path=None, attachment_name=None):
    """
    Send email with optional attachment
    """
    try:
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email]
        )
        
        if attachment_path and os.path.exists(attachment_path):
            email.attach_file(attachment_path, attachment_name)
        
        email.send()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def send_whatsapp_message(phone_number, message, media_url=None):
    """
    Send WhatsApp message (integrate with WhatsApp Business API)
    """
    # This would integrate with WhatsApp Business API
    # Implementation depends on your chosen WhatsApp service provider
    try:
        # Example implementation placeholder
        return True
    except Exception as e:
        print(f"Error sending WhatsApp: {e}")
        return False

def export_to_excel(data, filename, sheet_name='Sheet1'):
    """
    Export data to Excel format
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Set font style as per requirements (Bookman Old Style)
    font = Font(name='Bookman Old Style', size=10)
    
    # Headers
    if data and len(data) > 0:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(name='Bookman Old Style', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, (key, value) in enumerate(row_data.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = font
            
            # Format numbers and currency
            if isinstance(value, (int, float, Decimal)):
                if 'salary' in key.lower() or 'amount' in key.lower():
                    cell.number_format = '₹#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def convert_number_to_words(amount):
    """
    Convert number to words in Indian format
    """
    # This is a simplified version - you might want to use a library like num2words
    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
    teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    
    def convert_hundreds(n):
        result = ''
        if n >= 100:
            result += ones[n // 100] + ' Hundred '
            n %= 100
        if n >= 20:
            result += tens[n // 10] + ' '
            n %= 10
        elif n >= 10:
            result += teens[n - 10] + ' '
            n = 0
        if n > 0:
            result += ones[n] + ' '
        return result
    
    if amount == 0:
        return 'Zero Rupees Only'
    
    amount = int(amount)
    
    crores = amount // 10000000
    amount %= 10000000
    lakhs = amount // 100000
    amount %= 100000
    thousands = amount // 1000
    amount %= 1000
    hundreds = amount
    
    result = ''
    
    if crores:
        result += convert_hundreds(crores) + 'Crore '
    if lakhs:
        result += convert_hundreds(lakhs) + 'Lakh '
    if thousands:
        result += convert_hundreds(thousands) + 'Thousand '
    if hundreds:
        result += convert_hundreds(hundreds)
    
    return result.strip() + ' Rupees Only'

def format_name(name):
    """
    Format name as per requirement: First letter capital, rest small
    """
    if not name:
        return name
    
    # Split by spaces and capitalize first letter of each word
    words = name.lower().split()
    formatted_words = [word.capitalize() for word in words]
    return ' '.join(formatted_words)

def validate_date_format(date_string):
    """
    Validate date format as per requirement: DD/MM/YYYY
    """
    try:
        datetime.strptime(date_string, '%d/%m/%Y')
        return True
    except ValueError:
        return False

def format_date(date_obj):
    """
    Format date as per requirement: DD/MM/YYYY
    """
    if isinstance(date_obj, str):
        return date_obj
    return date_obj.strftime('%d/%m/%Y')

def generate_employee_code(company_prefix, sequence_number):
    """
    Generate employee code in format: COMPANY-000X
    """
    return f"{company_prefix}-{sequence_number:03d}"

def calculate_working_days(month, year, weekly_off_day=6):  # Saturday = 6
    """
    Calculate working days in a month excluding weekly offs and holidays
    """
    total_days = calendar.monthrange(year, month)[1]
    working_days = 0
    
    for day in range(1, total_days + 1):
        date_obj = date(year, month, day)
        if date_obj.weekday() != weekly_off_day:  # Exclude weekly off
            working_days += 1
    
    # Subtract standard holidays (you can customize this)
    # This is a simplified version - you might want to maintain a holidays table
    return working_days

def get_statutory_rates():
    """
    Get current statutory rates for PF, ESI, etc.
    """
    return {
        'pf_employee': 0.12,  # 12%
        'pf_employer': 0.13,  # 13%
        'esi_employee': 0.0175,  # 1.75%
        'esi_employer': 0.0325,  # 3.25%
        'esi_ceiling': 25000,  # ESI applicable up to 25,000
        'pt_rates': {
            'up_to_10000': 0,
            '10001_to_15000': 150,
            'above_15000': 200
        },
        'lwf_employee': 0.75,  # Fixed amount
        'lwf_employer': 1.25   # Fixed amount
    }

def calculate_pt_deduction(gross_salary):
    """
    Calculate PT deduction based on salary slabs
    """
    rates = get_statutory_rates()['pt_rates']
    
    if gross_salary <= 10000:
        return rates['up_to_10000']
    elif gross_salary <= 15000:
        return rates['10001_to_15000']
    else:
        return rates['above_15000']

def log_activity(user, action, description, model_name=None, object_id=None):
    """
    Log user activity for audit trail
    """
    # This would save to an ActivityLog model
    # Implementation depends on your activity logging requirements
    activity_data = {
        'user': user,
        'action': action,
        'description': description,
        'model_name': model_name,
        'object_id': object_id,
        'timestamp': datetime.now(),
        'ip_address': None  # You can get this from request
    }
    
    # Save to ActivityLog model
    # ActivityLog.objects.create(**activity_data)
    print(f"Activity logged: {user} - {action} - {description}")

def create_backup(model_instance, user):
    """
    Create backup before making changes to important data
    """
    # This would create a backup record
    # Implementation depends on your backup requirements
    pass

def validate_file_upload(file, allowed_extensions=None, max_size_mb=5):
    """
    Validate uploaded files
    """
    if allowed_extensions is None:
        allowed_extensions = ['pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx']
    
    # Check file extension
    file_extension = file.name.split('.')[-1].lower()
    if file_extension not in allowed_extensions:
        return False, f"File type '{file_extension}' not allowed"
    
    # Check file size
    if file.size > max_size_mb * 1024 * 1024:
        return False, f"File size exceeds {max_size_mb}MB limit"
    
    return True, "File is valid"